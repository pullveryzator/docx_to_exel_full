import os
import re

import pandas as pd
from openpyxl import load_workbook

from constants import TASK_SHEET_NAME, TOC_SHEET_NAME


def save_to_excel(data, output_file:str, sheet_name:str):
    """Сохранение данных в Excel с автоматическим удалением существующего листа."""
    df = pd.DataFrame(data)
    mode = 'a' if os.path.exists(output_file) else 'w'
    if mode == 'a':
        book = load_workbook(output_file)
        if sheet_name in book.sheetnames:
            book.remove(book[sheet_name])
        book.save(output_file)
    with pd.ExcelWriter(output_file, engine='openpyxl', mode=mode) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    return df


def reorder_sheets(output_file):
    """Переносит лист <TASK_SHEET_NAME> на первую позицию"""
    wb = load_workbook(output_file)
    if TASK_SHEET_NAME not in wb.sheetnames:
        print(f"Лист {TASK_SHEET_NAME} не найден в файле {output_file}")
        return None
    sheet = wb[TASK_SHEET_NAME]
    wb.remove(sheet)
    wb._sheets.insert(0, sheet)
    wb.save(output_file)


def excel_to_dict(excel_file: str):
    """Получение словаря из Excel файла,
    в котором ключи - это текст главы, а значения - столбец ID.
    Включена обработка ошибки на отсутствие нужного файла."""
    try:
        df = pd.read_excel(excel_file, sheet_name=TOC_SHEET_NAME)
        
        result_dict = dict(zip(df['name'], df['id']))
        
        return result_dict
        
    except FileNotFoundError:
        print(f"Файл {excel_file} не найден!")
        return None


def find_matching_paragraph(cleaned_text: str, toc: dict, trim_chars: int = 2) -> int:
    """
    Ищет наилучшее совпадение в словаре оглавления с возможностью обрезки символов.
    
    Параметры:
        cleaned_text - обработанный текст параграфа
        toc - словарь оглавления {название: id}
        trim_chars - количество символов для обрезки с конца (по умолчанию 2)
    """
    # Пробуем точное совпадение в первую очередь
    if cleaned_text in toc:
        return toc[cleaned_text]
    
    # Пробуем разные варианты обрезки
    for i in range(1, trim_chars + 1):
        truncated = cleaned_text[:-i] if len(cleaned_text) > i else cleaned_text
        for key in toc:
            if key.startswith(truncated):
                return toc[key]
    
    return None


def is_main_task(task_id: str) -> bool:
    """Проверяет, является ли идентификатор задачи основным номером.
    
    Основной номер задачи должен состоять из цифр и точки в конце (например, "5.", "72.").

    Args:
        task_id: Строка с идентификатором задачи для проверки

    Returns:
        True если идентификатор соответствует формату основного номера, иначе False

    Examples:
        >>> is_main_task("5.")
        True
        >>> is_main_task("5.1")
        False
    """
    return bool(re.fullmatch(r'^\d+\.$', str(task_id)))


def is_subtask(task_id: str, main_num: str) -> bool:
    """Проверяет, является ли задача подзадачей для указанного основного номера.

    Подзадача должна:
    1. Начинаться с основного номера (без точки на конце)
    2. Содержать после точки:
       - либо цифры (например, "5.1")
       - либо одну русскую букву (например, "5.а")

    Args:
        task_id: Идентификатор проверяемой задачи
        main_num: Основной номер задачи (может быть с точкой на конце)

    Returns:
        True если задача является подзадачей указанного основного номера, иначе False

    Examples:
        >>> is_subtask("5.1", "5.")
        True
        >>> is_subtask("5.а", "5")
        True
        >>> is_subtask("6.1", "5.")
        False
    """
    main_num = str(main_num).rstrip('.')
    pattern = rf'^{main_num}\.\d+$|^{main_num}\.[а-яё]$'
    return bool(re.fullmatch(pattern, str(task_id)))